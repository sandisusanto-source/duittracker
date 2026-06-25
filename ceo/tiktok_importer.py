"""
Importer khusus format export TikTok Shop Seller Center.

Mengenali & mengolah file native TikTok (tanpa perlu dirapikan manual):
  - "Semua Pesanan" (OrderSKUList)  -> sales_daily + products
  - "Campaign overview" (iklan harian) -> ad_spend_daily
  - "Product Traffic / Trend"        -> dikenali (info trafik, tidak diimpor)
  - "Income" (settlement & potongan) -> dikenali + ringkasan (tidak ditulis ke kas)
  - "Transaction Analysis"           -> dikenali (ringkasan bulanan)

Channel selalu = "TikTok Shop". Tanggal dinormalkan ke YYYY-MM-DD.
HPP TIDAK ada di export TikTok -> cost_price produk tidak disentuh
(diisi terpisah: manual atau via Accurate Online di fase lanjut).
"""

import io
import re
from datetime import datetime

from openpyxl import load_workbook

from db import get_conn

CHANNEL = "TikTok Shop"
CANCELLED = {"Dibatalkan", "Batal", "Canceled", "Cancelled"}


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9.\-]", "", str(v).replace("Rp", "").replace(".", "").replace(",", ""))
    try:
        return float(s) if s not in ("", "-", ".") else 0.0
    except ValueError:
        return 0.0


def _date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # ambil bagian tanggal saja
    s = s.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", str(v))
    if m:
        return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _rows(ws):
    return [r for r in ws.iter_rows(values_only=True)]


def detect_and_import(filename, content_bytes):
    """Kembalikan dict ringkasan kalau file dikenali sebagai TikTok, else None."""
    name = filename.lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        return None
    try:
        wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
    except Exception:
        return None

    sheet_titles = [ws.title for ws in wb.worksheets]

    # cari header di tiap sheet
    for ws in wb.worksheets:
        rows = _rows(ws)
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        hset = set(header)

        # --- Orders ("Semua Pesanan") ---
        if "Order ID" in hset and "Seller SKU" in hset:
            return _import_orders(rows, header, filename)

        # --- Campaign (iklan harian) ---
        if "Per Hari" in hset and "Biaya" in hset:
            return _import_campaign(rows, header, filename)

        # --- Transaction Analysis (ringkasan bulanan) ---
        if "GMV dari kreator" in hset:
            return {"ok": True, "type": "tiktok_transaction_analysis", "rows_imported": 0,
                    "filename": filename,
                    "note": "Ringkasan bulanan (GMV kreator/afiliasi) dikenali — tidak diimpor (pelengkap)."}

    # --- Product Traffic (header di baris ke-5, sheet 'Trend'/'Summary') ---
    for ws in wb.worksheets:
        rows = _rows(ws)
        for r in rows[:8]:
            cells = [str(c).strip() if c is not None else "" for c in r]
            if "Tanggal" in cells and "GMV" in cells and ("Impresi produk" in cells or "Pesanan SKU" in cells):
                return {"ok": True, "type": "tiktok_traffic", "rows_imported": 0, "filename": filename,
                        "note": "Data trafik harian (GMV, impresi, klik) dikenali — tidak diimpor (skema belum punya tabel trafik)."}

    # --- Income (settlement) ---
    if "Laporan" in sheet_titles or "Detail pesanan" in sheet_titles:
        return _summarize_income(wb, filename)

    return None


# ──────────────────────────────────────────────────────────────
def _import_orders(rows, header, filename):
    idx = {h: i for i, h in enumerate(header)}

    def g(r, name):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else None

    agg = {}            # (date, sku) -> [qty, revenue]
    orders_by_date = {}  # date -> set(order_id)
    prod = {}           # sku -> (name, category)

    for r in rows[1:]:
        oid = g(r, "Order ID")
        if not oid or str(oid).strip() in ("", "Platform unique order ID."):
            continue
        date = _date(g(r, "Created Time"))
        if not date:
            continue  # lewati baris deskripsi / tak bertanggal
        status = str(g(r, "Order Status") or "").strip()
        if status in CANCELLED:
            continue
        sku = str(g(r, "Seller SKU") or "").strip() or None
        qty = int(_num(g(r, "Quantity")))
        rev = _num(g(r, "SKU Subtotal After Discount"))
        if sku is None:
            continue
        a = agg.setdefault((date, sku), [0, 0.0])
        a[0] += qty
        a[1] += rev
        orders_by_date.setdefault(date, set()).add(str(oid))
        name = str(g(r, "Product Name") or sku).strip()
        cat = str(g(r, "Product Category") or "").strip() or None
        prod[sku] = (name, cat)

    n = 0
    with get_conn() as conn:
        # tulis master produk (jangan timpa cost_price/HPP yang sudah ada)
        for sku, (name, cat) in prod.items():
            conn.execute(
                """INSERT INTO products (sku, name, category) VALUES (?,?,?)
                   ON CONFLICT(sku) DO UPDATE SET name=excluded.name, category=excluded.category""",
                (sku, name, cat),
            )
        # tulis sales_daily; order_count distinct ditaruh di 1 baris per tanggal
        assigned = set()
        for (date, sku), (qty, rev) in agg.items():
            oc = 0
            if date not in assigned:
                oc = len(orders_by_date.get(date, ()))
                assigned.add(date)
            conn.execute(
                """INSERT INTO sales_daily (date, channel, sku, qty, revenue, order_count)
                   VALUES (?,?,?,?,?,?)
                   ON CONFLICT(date, channel, sku) DO UPDATE SET
                     qty=excluded.qty, revenue=excluded.revenue, order_count=excluded.order_count""",
                (date, CHANNEL, sku, qty, rev, oc),
            )
            n += 1

    total_orders = sum(len(s) for s in orders_by_date.values())
    return {"ok": True, "type": "tiktok_orders", "rows_imported": n, "filename": filename,
            "note": "%d order, %d SKU, %d hari (status batal dibuang)." % (
                total_orders, len(prod), len(orders_by_date))}


def _import_campaign(rows, header, filename):
    idx = {h: i for i, h in enumerate(header)}

    def col(name):
        return idx.get(name)

    i_date = col("Per Hari")
    i_cost = col("Biaya")
    i_conv = col("Pesanan SKU (Toko saat ini)")
    i_gmv = col("Penghasilan bruto (Toko saat ini)")

    n = 0
    with get_conn() as conn:
        for r in rows[1:]:
            raw = r[i_date] if i_date is not None else None
            date = _date(raw)
            if not date:
                continue  # lewati baris total ('-')
            cost = _num(r[i_cost]) if i_cost is not None else 0
            conv = int(_num(r[i_conv])) if i_conv is not None else 0
            gmv = _num(r[i_gmv]) if i_gmv is not None else 0
            conn.execute(
                """INSERT INTO ad_spend_daily
                   (date, channel, cost, impressions, clicks, conversions, revenue_attributed)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(date, channel) DO UPDATE SET
                     cost=excluded.cost, conversions=excluded.conversions,
                     revenue_attributed=excluded.revenue_attributed""",
                (date, CHANNEL, cost, 0, 0, conv, gmv),
            )
            n += 1
    return {"ok": True, "type": "tiktok_ads", "rows_imported": n, "filename": filename,
            "note": "Biaya & omset iklan harian TikTok diimpor."}


def _summarize_income(wb, filename):
    """Baca ringkasan 'Laporan' (tidak ditulis ke kas — cashflow butuh data internal)."""
    settlement = fees = revenue = None
    try:
        ws = wb["Laporan"]
        for r in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in r]
            label = " ".join(c for c in cells if c and not c.replace("-", "").replace(".", "").isdigit())
            val = None
            for c in reversed(cells):
                if c and re.fullmatch(r"-?\d+", c.replace(".", "")):
                    val = _num(c)
                    break
            if "penyelesaian pembayaran" in label.lower():
                settlement = val
            elif label.strip().startswith("Total Pendapatan"):
                revenue = val
            elif label.strip().startswith("Total Biaya"):
                fees = val
    except Exception:
        pass
    note = "Settlement (uang cair) dikenali"
    if settlement is not None:
        note += ": Rp{:,.0f}".format(settlement)
    note += ". Tidak ditulis ke Cashflow (perlu data kas internal Anda)."
    return {"ok": True, "type": "tiktok_income", "rows_imported": 0, "filename": filename, "note": note}
