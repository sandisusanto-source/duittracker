"""
Importer untuk export Accurate Online: "Kuantitas Barang per Gudang".

Memberi 2 hal sekaligus:
  - HPP per SKU  = Total Biaya / Kuantitas  -> products.cost_price
  - Stok per SKU = Kuantitas                -> inventory_snapshot

Kunci join = Kode Barang (Accurate) yang cocok dengan Seller SKU (TikTok).
Nama produk yang sudah ada (mis. dari TikTok) TIDAK ditimpa.
"""

import io
import re
from datetime import datetime

from openpyxl import load_workbook

from db import get_conn

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "mei": 5, "jun": 6,
    "jul": 7, "agu": 8, "agt": 8, "ags": 8, "sep": 9, "okt": 10, "nov": 11, "des": 12,
}


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("Rp", "").replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _report_date(rows):
    """Cari 'Per Tgl. 25 Jun 2026' di baris atas -> YYYY-MM-DD."""
    for r in rows[:8]:
        for c in r:
            if c and "tgl" in str(c).lower():
                m = re.search(r"(\d{1,2})\s+(\w{3})\w*\s+(\d{4})", str(c))
                if m:
                    mon = _MONTHS.get(m.group(2)[:3].lower())
                    if mon:
                        return "%04d-%02d-%02d" % (int(m.group(3)), mon, int(m.group(1)))
    return datetime.now().strftime("%Y-%m-%d")


def detect_and_import(filename, content_bytes):
    name = filename.lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        return None
    try:
        wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
    except Exception:
        return None

    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True)]

    # cari baris header yang punya "Kode Barang" + "Total Biaya"
    hidx = None
    for i, r in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c is not None else "" for c in r]
        if "kode barang" in cells and "total biaya" in cells:
            hidx = i
            break
    if hidx is None:
        wb.close()
        return None  # bukan format Accurate ini

    header = [str(c).strip().lower() if c is not None else "" for c in rows[hidx]]
    col = {name: header.index(name) for name in
           ("kode barang", "nama barang", "kuantitas", "total biaya") if name in header}
    c_kode = col.get("kode barang")
    c_qty = col.get("kuantitas")
    c_total = col.get("total biaya")
    c_name = col.get("nama barang")

    report_date = _report_date(rows)

    n_hpp = n_stock = 0
    with get_conn() as conn:
        for r in rows[hidx + 1:]:
            kode = r[c_kode] if c_kode is not None and c_kode < len(r) else None
            if not kode or not str(kode).strip():
                continue
            sku = str(kode).strip()
            qty = int(_num(r[c_qty])) if c_qty is not None else 0
            total = _num(r[c_total]) if c_total is not None else 0
            pname = (str(r[c_name]).strip() if c_name is not None and r[c_name] else sku)
            hpp = round(total / qty) if qty > 0 else 0

            # products: set HPP (jangan timpa nama yang sudah ada dari TikTok)
            conn.execute(
                """INSERT INTO products (sku, name, cost_price) VALUES (?,?,?)
                   ON CONFLICT(sku) DO UPDATE SET cost_price=excluded.cost_price""",
                (sku, pname, hpp),
            )
            if hpp > 0:
                n_hpp += 1

            # inventory snapshot (stok terkini)
            conn.execute(
                """INSERT INTO inventory_snapshot (date, sku, stock_qty) VALUES (?,?,?)
                   ON CONFLICT(date, sku) DO UPDATE SET stock_qty=excluded.stock_qty""",
                (report_date, sku, qty),
            )
            n_stock += 1
    wb.close()

    return {"ok": True, "type": "accurate_inventory", "rows_imported": n_stock,
            "filename": filename,
            "note": "HPP diisi untuk %d produk; stok %d SKU per %s." % (n_hpp, n_stock, report_date)}
