"""
Importer file XLSX/CSV untuk CEO Dashboard.

Sumber data MVP = upload file dari seller center / template.
Importer auto-deteksi JENIS file dari nama kolomnya, lalu menormalkan
ke tabel yang sesuai. Mendukung nama kolom Indonesia & Inggris.

Jenis file yang dikenali:
  - sales      : penjualan harian (date, channel, sku, qty, revenue, order)
  - ads        : biaya iklan harian (date, channel, cost, ...)
  - inventory  : snapshot stok (date, sku, stock)
  - cashflow   : kas masuk/keluar/piutang/hutang
  - products   : master produk + HPP
  - targets    : target omset bulanan
"""

import csv
import io
import re
from datetime import datetime

from db import get_conn

# ──────────────────────────────────────────────────────────────
# Alias kolom: petakan banyak nama header ke nama kanonik
# ──────────────────────────────────────────────────────────────
ALIASES = {
    "date": ["date", "tanggal", "tgl", "order date", "tanggal pesanan", "waktu"],
    "channel": ["channel", "marketplace", "toko", "platform"],
    "sku": ["sku", "kode produk", "product id", "id produk", "kode"],
    "product_name": ["product_name", "name", "nama produk", "nama", "product", "produk"],
    "category": ["category", "kategori"],
    "qty": ["qty", "quantity", "jumlah", "qty terjual", "terjual", "units"],
    "revenue": ["revenue", "omset", "omzet", "penjualan", "total", "sales", "nilai"],
    "order_count": ["order_count", "order", "orders", "jumlah order", "pesanan", "transaksi"],
    "cost": ["cost", "biaya iklan", "ad cost", "ad_cost", "biaya", "spend", "ad spend"],
    "impressions": ["impressions", "impresi", "tayangan"],
    "clicks": ["clicks", "klik"],
    "conversions": ["conversions", "konversi", "conv"],
    "revenue_attributed": ["revenue_attributed", "omset iklan", "penjualan iklan", "ad revenue", "gmv iklan"],
    "stock_qty": ["stock_qty", "stock", "stok", "sisa stok", "qty stok", "inventory", "persediaan"],
    "cost_price": ["cost_price", "hpp", "harga modal", "modal", "harga pokok"],
    "supplier": ["supplier", "pemasok", "vendor"],
    "type": ["type", "tipe", "jenis"],
    "amount": ["amount", "jumlah", "nilai", "nominal"],
    "note": ["note", "catatan", "keterangan", "ket"],
    "due_date": ["due_date", "jatuh tempo", "tempo"],
    "month": ["month", "bulan", "periode"],
    "revenue_target": ["revenue_target", "target", "target omset"],
}


def _canon(header):
    """Ubah satu nama header jadi nama kanonik (atau None)."""
    h = (header or "").strip().lower()
    for canon, names in ALIASES.items():
        if h in names:
            return canon
    return None


def _map_headers(headers):
    """Map list header -> {index: canon_name}."""
    mapping = {}
    for i, h in enumerate(headers):
        c = _canon(h)
        if c:
            mapping[i] = c
    return mapping


def _num(v):
    """Parse angka Indonesia: '5.200.000', 'Rp 5,2jt', '1.200' -> float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    if not s:
        return 0.0
    s = s.replace("rp", "").replace(" ", "")
    mult = 1
    if s.endswith("jt") or s.endswith("juta"):
        mult = 1_000_000
        s = re.sub(r"(jt|juta)$", "", s)
    elif s.endswith("rb") or s.endswith("ribu") or s.endswith("k"):
        mult = 1_000
        s = re.sub(r"(rb|ribu|k)$", "", s)
    # buang pemisah ribuan titik, ganti koma desimal jadi titik
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) * mult if s else 0.0
    except ValueError:
        return 0.0


def _date(v):
    """Normalisasi tanggal ke YYYY-MM-DD."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()[:19]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d",
                "%d/%m/%y", "%Y-%m-%d %H:%M:%S", "%d %b %Y"):
        try:
            return datetime.strptime(s.split(" ")[0] if fmt in ("%Y-%m-%d",) else s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # fallback: ambil pola YYYY-MM-DD jika ada
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


# ──────────────────────────────────────────────────────────────
# Baca file -> (headers, rows)
# ──────────────────────────────────────────────────────────────
def read_table(filename, content_bytes):
    """Baca CSV atau XLSX jadi (headers:list, rows:list[list])."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _read_xlsx(content_bytes)
    return _read_csv(content_bytes)


def _read_csv(content_bytes):
    text = content_bytes.decode("utf-8-sig", errors="replace")
    # deteksi delimiter (koma / titik koma)
    sample = text[:2048]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _read_xlsx(content_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row and any(c is not None and str(c).strip() for c in row):
            rows.append(list(row))
    wb.close()
    if not rows:
        return [], []
    return rows[0], rows[1:]


# ──────────────────────────────────────────────────────────────
# Deteksi jenis berdasarkan kolom kanonik yang ada
# ──────────────────────────────────────────────────────────────
def detect_type(canon_cols):
    cols = set(canon_cols)
    # 'type' (tipe/jenis) hanya muncul di file kas -> cek paling awal.
    # 'jumlah' bisa terpetakan ke qty maupun amount, jadi terima keduanya.
    if "type" in cols and ("amount" in cols or "qty" in cols):
        return "cashflow"
    if "cost_price" in cols and "sku" in cols and "revenue" not in cols:
        return "products"
    if "revenue_target" in cols or ("month" in cols and "revenue_target" in cols):
        return "targets"
    if "stock_qty" in cols and "revenue" not in cols:
        return "inventory"
    if "cost" in cols or "impressions" in cols or "revenue_attributed" in cols:
        return "ads"
    if "revenue" in cols or "qty" in cols or "order_count" in cols:
        return "sales"
    return None


# ──────────────────────────────────────────────────────────────
# Import per jenis
# ──────────────────────────────────────────────────────────────
def import_file(filename, content_bytes):
    """Entry point: baca, deteksi, simpan. Return ringkasan dict."""
    headers, rows = read_table(filename, content_bytes)
    if not headers:
        return {"ok": False, "error": "File kosong atau tidak terbaca."}

    mapping = _map_headers(headers)
    canon_cols = list(mapping.values())
    ftype = detect_type(canon_cols)
    if not ftype:
        return {"ok": False, "error": "Jenis file tidak dikenali. Kolom: %s" % headers}

    dict_rows = []
    for r in rows:
        d = {}
        for idx, canon in mapping.items():
            d[canon] = r[idx] if idx < len(r) else None
        dict_rows.append(d)

    handler = {
        "sales": _imp_sales,
        "ads": _imp_ads,
        "inventory": _imp_inventory,
        "cashflow": _imp_cashflow,
        "products": _imp_products,
        "targets": _imp_targets,
    }[ftype]

    count = handler(dict_rows)
    return {"ok": True, "type": ftype, "rows_imported": count, "filename": filename}


def _norm_channel(v):
    s = (str(v or "")).strip().lower()
    table = {
        "shopee": "Shopee", "shoppe": "Shopee",
        "tokopedia": "Tokopedia", "tokped": "Tokopedia",
        "tiktok": "TikTok Shop", "tiktok shop": "TikTok Shop", "tts": "TikTok Shop",
        "lazada": "Lazada", "laz": "Lazada",
    }
    return table.get(s, (str(v or "")).strip() or "Lainnya")


def _imp_sales(rows):
    n = 0
    with get_conn() as conn:
        for d in rows:
            date = _date(d.get("date"))
            channel = _norm_channel(d.get("channel"))
            if not date or not channel:
                continue
            sku = (str(d.get("sku")).strip() if d.get("sku") else None) or None
            conn.execute(
                """INSERT INTO sales_daily (date, channel, sku, qty, revenue, order_count)
                   VALUES (?,?,?,?,?,?)
                   ON CONFLICT(date, channel, sku) DO UPDATE SET
                     qty=excluded.qty, revenue=excluded.revenue, order_count=excluded.order_count""",
                (date, channel, sku, int(_num(d.get("qty"))),
                 _num(d.get("revenue")), int(_num(d.get("order_count")))),
            )
            # auto-buat produk minimal kalau ada nama
            if sku and d.get("product_name"):
                conn.execute(
                    "INSERT OR IGNORE INTO products (sku, name) VALUES (?,?)",
                    (sku, str(d.get("product_name")).strip()),
                )
            n += 1
    return n


def _imp_ads(rows):
    n = 0
    with get_conn() as conn:
        for d in rows:
            date = _date(d.get("date"))
            channel = _norm_channel(d.get("channel"))
            if not date or not channel:
                continue
            conn.execute(
                """INSERT INTO ad_spend_daily
                   (date, channel, cost, impressions, clicks, conversions, revenue_attributed)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(date, channel) DO UPDATE SET
                     cost=excluded.cost, impressions=excluded.impressions,
                     clicks=excluded.clicks, conversions=excluded.conversions,
                     revenue_attributed=excluded.revenue_attributed""",
                (date, channel, _num(d.get("cost")), int(_num(d.get("impressions"))),
                 int(_num(d.get("clicks"))), int(_num(d.get("conversions"))),
                 _num(d.get("revenue_attributed"))),
            )
            n += 1
    return n


def _imp_inventory(rows):
    n = 0
    with get_conn() as conn:
        for d in rows:
            date = _date(d.get("date")) or datetime.now().strftime("%Y-%m-%d")
            sku = str(d.get("sku")).strip() if d.get("sku") else None
            if not sku:
                continue
            conn.execute(
                """INSERT INTO inventory_snapshot (date, sku, stock_qty)
                   VALUES (?,?,?)
                   ON CONFLICT(date, sku) DO UPDATE SET stock_qty=excluded.stock_qty""",
                (date, sku, int(_num(d.get("stock_qty")))),
            )
            if d.get("product_name"):
                conn.execute(
                    "INSERT OR IGNORE INTO products (sku, name) VALUES (?,?)",
                    (sku, str(d.get("product_name")).strip()),
                )
            n += 1
    return n


def _imp_cashflow(rows):
    n = 0
    with get_conn() as conn:
        for d in rows:
            date = _date(d.get("date"))
            ttype = (str(d.get("type") or "")).strip().lower()
            if not date or ttype not in ("in", "out", "ar", "ap"):
                continue
            # 'jumlah' bisa terpetakan ke 'qty'; ambil amount dari mana pun yang terisi
            amount = d.get("amount")
            if amount is None or str(amount).strip() == "":
                amount = d.get("qty")
            conn.execute(
                """INSERT INTO cash_ledger (date, type, amount, category, note, due_date)
                   VALUES (?,?,?,?,?,?)""",
                (date, ttype, _num(amount), d.get("category"),
                 d.get("note"), _date(d.get("due_date"))),
            )
            n += 1
    return n


def _imp_products(rows):
    n = 0
    with get_conn() as conn:
        for d in rows:
            sku = str(d.get("sku")).strip() if d.get("sku") else None
            if not sku:
                continue
            conn.execute(
                """INSERT INTO products (sku, name, category, cost_price)
                   VALUES (?,?,?,?)
                   ON CONFLICT(sku) DO UPDATE SET
                     name=excluded.name, category=excluded.category,
                     cost_price=excluded.cost_price""",
                (sku, str(d.get("product_name") or sku).strip(),
                 d.get("category"), _num(d.get("cost_price"))),
            )
            n += 1
    return n


def _imp_targets(rows):
    n = 0
    with get_conn() as conn:
        for d in rows:
            month = str(d.get("month") or "").strip()[:7]
            if not re.match(r"\d{4}-\d{2}", month):
                continue
            channel = _norm_channel(d.get("channel")) if d.get("channel") else "ALL"
            conn.execute(
                """INSERT INTO targets (month, channel, revenue_target)
                   VALUES (?,?,?)
                   ON CONFLICT(month, channel) DO UPDATE SET
                     revenue_target=excluded.revenue_target""",
                (month, channel, _num(d.get("revenue_target"))),
            )
            n += 1
    return n
